# -*- coding: utf-8 -*-
"""
Build the video register as a polished Excel workbook.

    python video_register_xlsx.py

One row per module's video, read straight from content/areaNN/*.py — the
same dictionaries that build the decks and the web pages. Columns are laid
out to match the "Add Video or Document Link" form in the OneWork admin
console (Title / Content Type / Link) plus the extra detail a reviewer or
uploader needs, so a row can be copy-pasted straight into that form.
"""

import importlib
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

import build
import sitegen
import theme as T

ROOT = sitegen.ROOT
OUT_XLSX = os.path.join(ROOT, "docs", "Inducto-Video-Register.xlsx")

INK = "10182 6".replace(" ", "")
HEADER_FILL = PatternFill("solid", fgColor="101826")
HEADER_FONT = Font(name="Segoe UI Semibold", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Segoe UI", size=10)
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")
THIN = Side(style="thin", color="D9DEE7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def main():
    decks = [importlib.import_module(m).DECK for m in build.REGISTRY]
    order = {a: i for i, a in enumerate(sitegen.AREA_ORDER)}
    decks.sort(key=lambda d: (order[d["area"]], d["module_code"]))

    wb = Workbook()

    # ---- sheet 1: video register -------------------------------------
    ws = wb.active
    ws.title = "Video Register"

    cols = [
        ("Module Code", 12), ("Track", 22), ("Module Title", 30),
        ("Title", 34),                       # = OneWork "Title" field
        ("Content Type", 12),                # = OneWork "Content Type" (video)
        ("Link", 42),                        # = OneWork "Link (YouTube or any URL)"
        ("Channel", 26), ("Duration", 10), ("Duration (sec)", 12),
        ("Verified via", 30), ("Note to learner", 46),
    ]
    ws.append([c for c, _ in cols])
    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = BORDER
    ws.freeze_panes = "A2"

    def to_seconds(dur):
        parts = [int(p) for p in dur.split(":")]
        while len(parts) < 3:
            parts.insert(0, 0)
        h, m, s = parts
        return h * 3600 + m * 60 + s

    for d in decks:
        v = d.get("video")
        row = [
            d["module_code"], T.AREAS[d["area"]]["name"], d["title"],
            v["title"] if v else "", "video" if v else "",
            v["url"] if v else "", v["channel"] if v else "",
            v["duration"] if v else "", to_seconds(v["duration"]) if v else "",
            "YouTube oEmbed + watch-page runtime, checked before writing "
            "into the module" if v else "No video attached — see note",
            v["note"] if v else "This module has no video. Nothing was "
            "substituted for it.",
        ]
        ws.append(row)

    last = ws.max_row
    for r in range(2, last + 1):
        for c in range(1, len(cols) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = WRAP if c in (4, 11) else TOP

    tbl = Table(displayName="VideoRegister", ref="A1:%s%d" % (get_column_letter(len(cols)), last))
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    # ---- sheet 2: how to use this -------------------------------------
    ws2 = wb.create_sheet("How to use this")
    ws2.column_dimensions["A"].width = 100
    notes = [
        ("Inducto Learning & Knowledge Library — video register", True),
        ("", False),
        ("One row per module's video. Every title, channel and runtime was "
         "read back from YouTube itself (oEmbed + the watch page), not "
         "copied from a list, so what is in this sheet is what is actually "
         "on YouTube.", False),
        ("", False),
        ("To add these to the OneWork admin console:", True),
        ("1. Training → open the module → \U0001F3AC Videos & documents → + Add Link.", False),
        ("2. Title = the Title column. Content Type = Video. "
         "Link = the Link column.", False),
        ("3. Repeat for each module. No file upload is needed — OneWork "
         "stores the YouTube link and embeds it.", False),
        ("", False),
        ("These are third-party videos, not company material. Where a video "
         "differs from the module text, the module text is correct — each "
         "module page says so.", False),
    ]
    for i, (text, bold) in enumerate(notes, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        cell.font = Font(name="Segoe UI Semibold" if bold else "Segoe UI",
                         size=13 if bold and i == 1 else 10.5, bold=bold,
                         color="101826")
        cell.alignment = WRAP

    wb.save(OUT_XLSX)
    with_video = sum(1 for d in decks if d.get("video"))
    print("Saved %s" % OUT_XLSX)
    print("  %d rows, %d with a video, %d without"
         % (len(decks), with_video, len(decks) - with_video))


if __name__ == "__main__":
    main()
